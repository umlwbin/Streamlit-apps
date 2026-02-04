import streamlit as st
import zipfile
from zipfile import ZipFile
import os
import io
import sys
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

sys.path.append(f'{os.path.abspath(os.curdir)}/Modules')
import session_initializer


def download_output():

    def on_download():
        st.balloons()
        session_initializer.reset_widget_flags()

    current_files = st.session_state.current_data
    if not current_files:
        st.info("No processed files available yet.")
        return

    st.markdown("##### 📑 CSV")

    # Detect merged output safely
    merged_filename = None
    for name in current_files.keys():
        if name.lower().startswith("merged"):
            merged_filename = name
            break

    # CASE 1 — Merged file
    if merged_filename:
        df = current_files[merged_filename]
        st.download_button(
            label=f"⬇️ Download {merged_filename}",
            data=df.to_csv(index=False).encode("utf-8"),
            file_name=merged_filename,
            mime="text/csv",
            on_click=on_download,
            icon=":material/download:"
        )
        return

    # CASE 2 — Single file
    if len(current_files) == 1:
        filename, df = next(iter(current_files.items()))
        base, ext = os.path.splitext(filename)

        st.download_button(
            label=f"⬇️ Download {filename}",
            data=df.to_csv(index=False).encode("utf-8"),
            file_name=f"{base}_cleaned.csv",
            mime="text/csv",
            on_click=on_download,
            icon=":material/download:"
        )
        return

    # CASE 3 — Multiple files → ZIP
    zip_buffer = io.BytesIO()
    with ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # Main cleaned files
        for filename, df in current_files.items():
            base, ext = os.path.splitext(filename)
            csv_bytes = df.to_csv(index=False).encode("utf-8")
            zip_file.writestr(f"{base}_cleaned.csv", csv_bytes)

        # Supplementary outputs (e.g., RVQ summaries)
        if "supplementary_outputs" in st.session_state:
            for summary_name, summary_df in st.session_state.supplementary_outputs.items():
                zip_file.writestr(summary_name, summary_df.to_csv(index=False))


    zip_buffer.seek(0)

    st.download_button(
        label="⬇️ Download All as ZIP",
        data=zip_buffer,
        file_name="cleaned_files.zip",
        mime="application/zip",
        on_click=on_download,
        icon=":material/download:"
    )

# EXCEL-------------------------------------------------------------------------------------------------
def to_excel_with_formatting(df, freeze_header=False):

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"

    # Identify datetime columns
    datetime_cols = [col for col in df.columns if pd.api.types.is_datetime64_any_dtype(df[col])]

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    datetime_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # soft yellow

    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # Write data rows
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            if pd.isna(cell_value): # Convert <NA> to None
                cell_value = None
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            if df.columns[col_num - 1] in datetime_cols:
                cell.fill = datetime_fill

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Freeze header row if requested
    if freeze_header:
        ws.freeze_panes = "A2"

    wb.save(output)
    output.seek(0)
    return output

def excel_download():
    if len(st.session_state.current_data) == 1:
        filename, df = next(iter(st.session_state.current_data.items()))
        base, ext = os.path.splitext(filename)

        st.markdown(" ")
        st.markdown("###### 📑 EXCEL")
        freeze = st.checkbox("Freeze header row in Excel")

        excel_data = to_excel_with_formatting(df, freeze_header=freeze)

        st.download_button(
            label="⬇️ Download Excel File with Formatting",
            data=excel_data,
            file_name=f"{base}_cleaned.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            icon=":material/download:"
        )