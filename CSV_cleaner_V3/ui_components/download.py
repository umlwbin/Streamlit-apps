import streamlit as st
import pandas as pd
import io
import os
from zipfile import ZipFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import state.session_initializer as session_initializer


# ---------------------------------------------------------
# Helper: Reset widget flags + balloons
# ---------------------------------------------------------
def _on_download():
    st.balloons()
    session_initializer.reset_widget_flags()


# ---------------------------------------------------------
# CSV DOWNLOADS
# ---------------------------------------------------------
def download_output():
    """
    Handles CSV download logic for:
        - merged output
        - single cleaned file
        - multiple cleaned files (ZIP)
    """

    current_files = st.session_state.current_data

    if not current_files:
        st.info("No processed files available yet.")
        return

    st.markdown("##### 📑 CSV")

    # -----------------------------------------------------
    # Detect merged output (one-liner)
    # -----------------------------------------------------
    merged_name = next((name for name in current_files if name.lower().startswith("merged")), None)

    # -----------------------------------------------------
    # CASE 1 - Merged file
    # -----------------------------------------------------
    if merged_name:
        df = current_files[merged_name]
        st.download_button(
            label=f"⬇️ Download {merged_name}",
            data=df.to_csv(index=False).encode("utf-8"),
            file_name=merged_name,
            mime="text/csv",
            on_click=_on_download,
            icon=":material/download:"
        )
        return

    # -----------------------------------------------------
    # CASE 2 - Single file
    # -----------------------------------------------------
    if len(current_files) == 1:
        filename, df = next(iter(current_files.items()))
        base, _ = os.path.splitext(filename)

        st.download_button(
            label=f"⬇️ Download {filename}",
            data=df.to_csv(index=False).encode("utf-8"),
            file_name=f"{base}_cleaned.csv",
            mime="text/csv",
            on_click=_on_download,
            icon=":material/download:"
        )
        return

    # -----------------------------------------------------
    # CASE 3 - Multiple files --> ZIP
    # -----------------------------------------------------
    zip_buffer = io.BytesIO()

    with ZipFile(zip_buffer, "w") as zip_file:

        # Add cleaned CSVs
        for filename, df in current_files.items():
            base, _ = os.path.splitext(filename)
            zip_file.writestr(
                f"{base}_cleaned.csv",
                df.to_csv(index=False).encode("utf-8")
            )

        # Add supplementary outputs (e.g., RVQ metadata)
        supp = st.session_state.get("supplementary_outputs", {})
        for name, df in supp.items():
            zip_file.writestr(name, df.to_csv(index=False))

    zip_buffer.seek(0)

    st.download_button(
        label="⬇️ Download All as ZIP",
        data=zip_buffer,
        file_name="cleaned_files.zip",
        mime="application/zip",
        on_click=_on_download,
        icon=":material/download:"
    )


# ---------------------------------------------------------
# EXCEL DOWNLOADS
# ---------------------------------------------------------
def to_excel_with_formatting(df, freeze_header=False):
    """
    Convert a DataFrame to an Excel file with:
        - bold headers
        - shaded header row
        - highlighted datetime columns
        - auto column widths
        - optional frozen header row
    """

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"

    # Identify datetime columns
    datetime_cols = [
        col for col in df.columns
        if pd.api.types.is_datetime64_any_dtype(df[col])
    ]

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    datetime_fill = PatternFill("solid", fgColor="FFF2CC")

    # Write headers
    for col_num, title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    # Write data rows
    for row_num, row in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if df.columns[col_num - 1] in datetime_cols:
                cell.fill = datetime_fill

    # Auto column widths
    for column_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

    # Freeze header row
    if freeze_header:
        ws.freeze_panes = "A2"

    wb.save(output)
    output.seek(0)
    return output


def excel_download():
    """
    Excel download is only available when exactly one file exists.
    """
    if len(st.session_state.current_data) != 1:
        return

    filename, df = next(iter(st.session_state.current_data.items()))
    base, _ = os.path.splitext(filename)

    st.markdown(" ")
    st.markdown("###### 📑 EXCEL")

    freeze = st.checkbox("Freeze header row in Excel")

    excel_data = to_excel_with_formatting(df, freeze_header=freeze)

    st.download_button(
        label="⬇️ Download Excel File with Formatting",
        data=excel_data,
        file_name=f"{base}_cleaned.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        icon=":material/download:"
    )
